from django.conf import settings
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
import json
import openpyxl
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile
import re

from django.contrib.admin.views.decorators import staff_member_required
import xlwt
from datetime import timedelta

from io import StringIO

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import black, grey, HexColor
from PIL import Image as PILImage
import qrcode
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import tempfile


from .models import (
    CustomUser, Subscription, SubscriptionPlan, City, Zone,
    SubscriptionQRCode, CollectionRequest, Payment, Tricycle
)

# Vérifier si l'utilisateur est admin
def is_admin(user):
    return user.is_authenticated and user.user_type == 'admin'

@login_required(login_url='login')
@user_passes_test(is_admin)
def gestion_abonnements(request):
    """Page principale de gestion des abonnements et utilisateurs"""
    # Récupérer les paramètres de filtrage
    statut_filter = request.GET.get('statut', 'all')
    search_query = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)
    
    # Filtrage des abonnements
    subscriptions = Subscription.objects.select_related(
        'user', 'plan', 'zone', 'address'
    ).prefetch_related('collection_days')
    
    if statut_filter != 'all':
        subscriptions = subscriptions.filter(status=statut_filter)
    
    if search_query:
        subscriptions = subscriptions.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__phone__icontains=search_query) |
            Q(plan__name__icontains=search_query)
        )
    
    # Filtrage des utilisateurs
    users = CustomUser.objects.select_related('city')
    if search_query:
        users = users.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Pagination
    subscription_paginator = Paginator(subscriptions.order_by('-created_at'), 10)
    users_paginator = Paginator(users.order_by('-date_joined'), 10)
    
    subscriptions_page = subscription_paginator.get_page(page_number)
    users_page = users_paginator.get_page(page_number)
    
    # Statistiques
    total_subscriptions = subscriptions.count()
    active_subscriptions = subscriptions.filter(status='active').count()
    suspended_subscriptions = subscriptions.filter(status='suspended').count()
    total_users = users.count()
    
    context = {
        'users': users_page,
        'subscriptions': subscriptions_page,
        'subscription_plans': SubscriptionPlan.objects.filter(is_active=True),
        'cities': City.objects.all(),
        'zones': Zone.objects.filter(is_active=True),
        'total_subscriptions': total_subscriptions,
        'active_subscriptions': active_subscriptions,
        'suspended_subscriptions': suspended_subscriptions,
        'total_users': total_users,
        'current_filter': statut_filter,
        'search_query': search_query,
    }
    
    return render(request, 'administrations/gestion_abonnements.html', context)




    

@staff_member_required
def exporter_tous_qrcodes_pdf(request):
    """
    Exporte tous les QR codes des abonnements dans un fichier PDF format A4
    avec 6 QR codes par page, classés par ordre alphabétique des utilisateurs
    """
    # Récupérer tous les abonnements actifs avec leurs QR codes
    subscriptions = Subscription.objects.filter(
        qr_code__isnull=False
    ).select_related('user', 'plan', 'qr_code').order_by('user__last_name', 'user__first_name')
    
    if not subscriptions.exists():
        # Si aucun abonnement avec QR code, essayer de récupérer ceux qui ont un QR code via l'OneToOne
        subscriptions = Subscription.objects.filter(
            status='active',
            qr_code__isnull=False
        ).select_related('user', 'plan', 'qr_code').order_by('user__last_name', 'user__first_name')
    
    # Créer la réponse HTTP avec le type PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tous_les_qrcodes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Créer le canvas PDF avec dimensions A4
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4  # 595.27 x 841.89 points (environ 21cm x 29.7cm)
    
    # Paramètres de mise en page
    margin = 1.5 * cm
    qr_width = (width - 2 * margin) / 3  # 3 QR codes par ligne
    qr_height = (height - 2 * margin) / 2  # 2 lignes par page
    
    # Chemin du logo (à adapter selon votre structure)
    logo_path = None
    # Essayez de localiser le logo dans différents emplacements possibles
    possible_logo_paths = [
        os.path.join(settings.MEDIA_ROOT, 'logo.png'),
        os.path.join(settings.STATIC_ROOT, 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png'),
    ]
    
    for path in possible_logo_paths:
        if os.path.exists(path):
            logo_path = path
            break
    
    # Si aucun logo trouvé, on utilisera un texte
    logo_found = logo_path is not None
    
    # Fonction pour générer l'image QR code
    def generate_qr_image(data, size=200):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        return img
    
    # Compter le nombre total de QR codes
    total_qr_codes = subscriptions.count()
    current_page = 1
    qr_index = 0
    
    # Créer un dictionnaire pour stocker temporairement les images
    temp_images = []
    
    try:
        # Pour chaque abonnement, préparer l'image QR
        for subscription in subscriptions:
            # Récupérer les informations pour le QR code
            user = subscription.user
            user_name = user.get_full_name() or user.username
            user_phone = user.phone or "Non renseigné"
            
            # Données à encoder dans le QR code
            if hasattr(subscription, 'qr_code') and subscription.qr_code:
                # Utiliser l'URL de renouvellement existante
                qr_data = subscription.qr_code.get_renewal_url()
            else:
                # Créer une URL par défaut
                qr_data = f"{settings.SITE_URL}/subscription/{subscription.id}/"
            
            # Générer l'image QR
            qr_img = generate_qr_image(qr_data, size=300)
            
            # Sauvegarder temporairement l'image
            img_buffer = BytesIO()
            qr_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Ouvrir avec PIL pour redimensionner si nécessaire
            pil_img = PILImage.open(img_buffer)
            
            # Convertir en mode RGB si nécessaire
            if pil_img.mode != 'RGB':
                pil_img = pil_img.convert('RGB')
            
            # Sauvegarder temporairement
            temp_img = BytesIO()
            pil_img.save(temp_img, format='PNG')
            temp_img.seek(0)
            
            temp_images.append({
                'img': temp_img,
                'user_name': user_name,
                'user_phone': user_phone,
                'user': user
            })
        
        # Trier les images par nom d'utilisateur (déjà trié par la requête)
        # Mais on garde le tri au cas où
        
        # Parcourir toutes les images et les placer sur les pages
        for idx, temp_data in enumerate(temp_images):
            # Calculer la position sur la page
            col = idx % 3
            row = (idx // 3) % 2
            
            x = margin + col * qr_width
            y = height - margin - (row + 1) * qr_height + 0.5*cm  # Ajustement
            
            # Si on commence une nouvelle page (après 6 QR codes)
            if idx > 0 and idx % 6 == 0:
                p.showPage()  # Nouvelle page
                current_page += 1
            
            # Dessiner le cadre du QR code
            p.setStrokeColor(grey)
            p.setLineWidth(0.5)
            p.rect(x, y - qr_height + 1*cm, qr_width - 0.5*cm, qr_height - 1*cm, stroke=1, fill=0)
            
            # Ajouter le logo en haut du cadre
            if logo_found:
                try:
                    logo = ImageReader(logo_path)
                    logo_width = 2.5*cm
                    logo_height = 2.5*cm
                    logo_x = x + (qr_width - 0.5*cm - logo_width) / 2
                    logo_y = y - 0.5*cm
                    p.drawImage(logo, logo_x, logo_y - logo_height + 0.3*cm, 
                                width=logo_width, height=logo_height, 
                                preserveAspectRatio=True, mask='auto')
                except Exception as e:
                    # En cas d'erreur avec le logo, écrire du texte
                    p.setFont("Helvetica-Bold", 12)
                    p.setFillColor(HexColor("#7B1212"))  # Vert ECOCITY
                    p.drawCentredString(x + (qr_width - 0.5*cm) / 2, y - 0.3*cm, "ECOCITY")
                    p.setFillColor(black)
            else:
                # Pas de logo, écrire le nom de l'application
                p.setFont("Helvetica-Bold", 12)
                p.setFillColor(HexColor("#F45A02"))  # Vert ECOCITY
                p.drawCentredString(x + (qr_width - 0.5*cm) / 2, y - 0.3*cm, "ECOCITY")
                p.setFillColor(black)
            
            # Charger l'image depuis le buffer temporaire
            temp_data['img'].seek(0)
            img_reader = ImageReader(temp_data['img'])
            
            # Dimensions du QR code dans le PDF
            qr_img_width = 4*cm
            qr_img_height = 4*cm
            qr_img_x = x + (qr_width - 0.5*cm - qr_img_width) / 2
            qr_img_y = y - 2*cm - qr_img_height
            
            # Dessiner le QR code
            p.drawImage(img_reader, qr_img_x, qr_img_y, 
                       width=qr_img_width, height=qr_img_height, 
                       preserveAspectRatio=True, mask='auto')
            
            # Ajouter les informations utilisateur en bas
            p.setFont("Helvetica", 8)
            p.setFillColor(black)
            
            # Nom de l'utilisateur
            user_name_display = temp_data['user_name']
            if len(user_name_display) > 25:
                user_name_display = user_name_display[:22] + "..."
            
            p.drawCentredString(x + (qr_width - 0.5*cm) / 2, qr_img_y - 0.5*cm, 
                               f"{user_name_display}")
            
            # Téléphone
            phone_display = f"Tel: {temp_data['user_phone']}"
            if len(phone_display) > 20:
                phone_display = phone_display[:17] + "..."
            
            p.drawCentredString(x + (qr_width - 0.5*cm) / 2, qr_img_y - 1*cm, 
                               phone_display)
            
            
            # Optionnel : ajouter le type de plan ou l'ID
            if hasattr(subscription, 'plan') and subscription.plan:
                plan_display = subscription.plan.name
                if len(plan_display) > 15:
                    plan_display = plan_display[:12] + "..."
                p.setFont("Helvetica", 6)
                p.drawCentredString(x + (qr_width - 0.5*cm) / 2, qr_img_y - 1.3*cm, 
                                   plan_display)
        
        # Ajouter un numéro de page
        p.setFont("Helvetica", 8)
        p.setFillColor(grey)
        p.drawRightString(width - margin, margin/2, f"Page {current_page}")
        
        p.save()
        
        # Récupérer le contenu du PDF
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
    finally:
        # Nettoyer les buffers temporaires
        for temp_data in temp_images:
            temp_data['img'].close()
    
    return response


@staff_member_required
def exporter_qrcodes_pdf_par_zone(request, zone_id=None):
    """
    Exporte les QR codes des abonnements par zone géographique
    """
    from .models import Zone
    
    if zone_id:
        zone = get_object_or_404(Zone, id=zone_id)
        subscriptions = Subscription.objects.filter(
            status='active',
            zone=zone,
            qr_code__isnull=False
        ).select_related('user', 'plan', 'qr_code').order_by('user__last_name', 'user__first_name')
        
        filename = f"qrcodes_zone_{zone.nom}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    else:
        # Si pas de zone spécifiée, exporter toutes les zones
        subscriptions = Subscription.objects.filter(
            status='active',
            qr_code__isnull=False
        ).select_related('user', 'plan', 'qr_code', 'zone').order_by('zone__nom', 'user__last_name', 'user__first_name')
        
        filename = f"tous_les_qrcodes_par_zone_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    # Même logique que la fonction précédente pour la génération du PDF
    # (vous pouvez copier le code de exporter_tous_qrcodes_pdf ici)
    # Pour éviter la duplication, nous allons appeler une fonction commune
    
    return generate_qrcodes_pdf_response(subscriptions, filename)


def generate_qrcodes_pdf_response(subscriptions, filename):
    """
    Fonction commune pour générer un PDF de QR codes
    """
    if not subscriptions.exists():
        response = HttpResponse(content_type='text/plain')
        response.write("Aucun QR code trouvé pour les abonnements sélectionnés.")
        return response
    
    # Créer la réponse HTTP avec le type PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Créer le canvas PDF avec dimensions A4
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Paramètres de mise en page (mêmes que précédemment)
    margin = 1.5 * cm
    qr_width = (width - 2 * margin) / 3
    qr_height = (height - 2 * margin) / 2
    
    # Recherche du logo
    logo_path = None
    possible_logo_paths = [
        os.path.join(settings.MEDIA_ROOT, 'logo.png'),
        os.path.join(settings.STATIC_ROOT, 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png'),
    ]
    
    for path in possible_logo_paths:
        if os.path.exists(path):
            logo_path = path
            break
    
    logo_found = logo_path is not None
    
    # Fonction pour générer l'image QR code
    def generate_qr_image(data, size=200):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        return img
    
    temp_images = []
    
    try:
        for subscription in subscriptions:
            user = subscription.user
            user_name = user.get_full_name() or user.username
            user_phone = user.phone or "Non renseigné"
            
            if hasattr(subscription, 'qr_code') and subscription.qr_code:
                qr_data = subscription.qr_code.get_renewal_url()
            else:
                qr_data = f"{settings.SITE_URL}/subscription/{subscription.id}/"
            
            qr_img = generate_qr_image(qr_data, size=300)
            
            img_buffer = BytesIO()
            qr_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            pil_img = PILImage.open(img_buffer)
            if pil_img.mode != 'RGB':
                pil_img = pil_img.convert('RGB')
            
            temp_img = BytesIO()
            pil_img.save(temp_img, format='PNG')
            temp_img.seek(0)
            
            temp_images.append({
                'img': temp_img,
                'user_name': user_name,
                'user_phone': user_phone,
                'zone_name': subscription.zone.nom if subscription.zone else "Sans zone"
            })
        
        # Grouper par zone si nécessaire pour l'affichage
        current_zone = None
        
        for idx, temp_data in enumerate(temp_images):
            col = idx % 3
            row = (idx // 3) % 2
            
            x = margin + col * qr_width
            y = height - margin - (row + 1) * qr_height + 0.5*cm
            
            if idx > 0 and idx % 6 == 0:
                p.showPage()
            
            # Ajouter le nom de la zone si elle change
            if temp_data['zone_name'] != current_zone and idx % 6 == 0:
                current_zone = temp_data['zone_name']
                p.setFont("Helvetica-Bold", 10)
                p.setFillColor(HexColor("#F06C0D"))
                p.drawString(margin, height - margin - 0.3*cm, f"Zone: {current_zone}")
                p.setFillColor(black)
            
            # Dessiner le cadre
            p.setStrokeColor(grey)
            p.setLineWidth(0.5)
            p.rect(x, y - qr_height + 1*cm, qr_width - 0.5*cm, qr_height - 1*cm, stroke=1, fill=0)
            
            # Ajouter le logo
            if logo_found:
                try:
                    logo = ImageReader(logo_path)
                    logo_width = 2*cm
                    logo_height = 0.8*cm
                    logo_x = x + (qr_width - 0.5*cm - logo_width) / 2
                    logo_y = y - 0.5*cm
                    p.drawImage(logo, logo_x, logo_y - logo_height + 0.3*cm, 
                               width=logo_width, height=logo_height, 
                               preserveAspectRatio=True, mask='auto')
                except:
                    p.setFont("Helvetica-Bold", 12)
                    p.setFillColor(HexColor("#2E8B57"))
                    p.drawCentredString(x + (qr_width - 0.5*cm) / 2, y - 0.3*cm, "ECOCITY")
                    p.setFillColor(black)
            else:
                p.setFont("Helvetica-Bold", 12)
                p.setFillColor(HexColor("#2E8B57"))
                p.drawCentredString(x + (qr_width - 0.5*cm) / 2, y - 0.3*cm, "ECOCITY")
                p.setFillColor(black)
            
            temp_data['img'].seek(0)
            img_reader = ImageReader(temp_data['img'])
            
            qr_img_width = 4*cm
            qr_img_height = 4*cm
            qr_img_x = x + (qr_width - 0.5*cm - qr_img_width) / 2
            qr_img_y = y - 2*cm - qr_img_height
            
            p.drawImage(img_reader, qr_img_x, qr_img_y, 
                       width=qr_img_width, height=qr_img_height, 
                       preserveAspectRatio=True, mask='auto')
            
            p.setFont("Helvetica", 8)
            p.setFillColor(black)
            
            user_name_display = temp_data['user_name']
            if len(user_name_display) > 25:
                user_name_display = user_name_display[:22] + "..."
            
            p.drawCentredString(x + (qr_width - 0.5*cm) / 2, qr_img_y - 0.5*cm, 
                               f"{user_name_display}")
            
            phone_display = f"Tel: {temp_data['user_phone']}"
            if len(phone_display) > 20:
                phone_display = phone_display[:17] + "..."
            
            p.drawCentredString(x + (qr_width - 0.5*cm) / 2, qr_img_y - 1*cm, 
                               phone_display)
        
        p.setFont("Helvetica", 8)
        p.setFillColor(grey)
        p.drawRightString(width - margin, margin/2, f"Page {p.getPageNumber()}")
        
        p.save()
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
    finally:
        for temp_data in temp_images:
            temp_data['img'].close()
    
    return response

@login_required(login_url='login')
@user_passes_test(is_admin)
def ajouter_utilisateur(request):
    """Ajouter un nouvel utilisateur"""
    if request.method == 'POST':
        try:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            user_type = request.POST.get('user_type')
            city_id = request.POST.get('city')
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            # Validation
            if not all([first_name, last_name, email, user_type]):
                messages.error(request, "Tous les champs obligatoires doivent être remplis.")
                return redirect('gestion_abonnements')
            
            if password != confirm_password:
                messages.error(request, "Les mots de passe ne correspondent pas.")
                return redirect('gestion_abonnements')
            
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, "Un utilisateur avec cet email existe déjà.")
                return redirect('gestion_abonnements')
            
            # Création de l'utilisateur
            user = CustomUser(
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                user_type=user_type,
                username=email  # Utiliser l'email comme nom d'utilisateur
            )
            
            if city_id:
                user.city_id = city_id
            
            user.set_password(password)
            user.save()
            
            messages.success(request, f"Utilisateur {user.get_full_name()} créé avec succès.")
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création de l'utilisateur: {str(e)}")
    
    return redirect('gestion_abonnements')

@login_required(login_url='login')
@user_passes_test(is_admin)
def editer_utilisateur(request):
    """Éditer un utilisateur existant"""
    if request.method == 'POST':
        try:
            user_id = request.POST.get('user_id')
            user = get_object_or_404(CustomUser, id=user_id)
            
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.phone = request.POST.get('phone')
            user.user_type = request.POST.get('user_type')
            user.city_id = request.POST.get('city') or None
            user.is_verified = request.POST.get('is_verified') == 'on'
            
            # Gestion du mot de passe
            password = request.POST.get('password')
            if password:
                user.set_password(password)
            
            user.save()
            
            messages.success(request, f"Utilisateur {user.get_full_name()} mis à jour avec succès.")
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour de l'utilisateur: {str(e)}")
    
    return redirect('gestion_abonnements')

@login_required(login_url='login')
@user_passes_test(is_admin)
def supprimer_utilisateur(request, user_id):
    """Supprimer un utilisateur"""
    if request.method == 'POST':
        try:
            user = get_object_or_404(CustomUser, id=user_id)
            
            # Vérifier s'il y a des abonnements associés
            if user.subscriptions.exists():
                messages.error(request, "Impossible de supprimer cet utilisateur car il a des abonnements actifs.")
                return redirect('gestion_abonnements')
            
            user_name = user.get_full_name()
            user.delete()
            
            messages.success(request, f"Utilisateur {user_name} supprimé avec succès.")
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression de l'utilisateur: {str(e)}")
    
    return redirect('gestion_abonnements')

@login_required(login_url='login')
@user_passes_test(is_admin)
def editer_abonnement(request):
    """Éditer un abonnement existant"""
    if request.method == 'POST':
        try:
            subscription_id = request.POST.get('subscription_id')
            subscription = get_object_or_404(Subscription, id=subscription_id)
            
            subscription.user_id = request.POST.get('user')
            subscription.plan_id = request.POST.get('plan')
            subscription.status = request.POST.get('status')
            subscription.start_date = request.POST.get('start_date')
            subscription.end_date = request.POST.get('end_date') or None
            subscription.custom_price = request.POST.get('custom_price') or None
            subscription.special_instructions = request.POST.get('special_instructions', '')
            
            subscription.save()
            
            messages.success(request, f"Abonnement {subscription.plan.name} mis à jour avec succès.")
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour de l'abonnement: {str(e)}")
    
    return redirect('gestion_abonnements')

@login_required(login_url='login')
@user_passes_test(is_admin)
def supprimer_abonnement(request, subscription_id):
    """Supprimer un abonnement"""
    if request.method == 'POST':
        try:
            subscription = get_object_or_404(Subscription, id=subscription_id)
            subscription_name = f"{subscription.plan.name} - {subscription.user.get_full_name()}"
            subscription.delete()
            
            messages.success(request, f"Abonnement {subscription_name} supprimé avec succès.")
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression de l'abonnement: {str(e)}")
    
    return redirect('gestion_abonnements')

@login_required(login_url='login')
@user_passes_test(is_admin)
def generer_qr_code(request, subscription_id):
    """Générer ou récupérer un QR Code pour un abonnement"""
    try:
        subscription = get_object_or_404(Subscription, id=subscription_id)
        
        # Vérifier si un QR Code existe déjà
        qr_code, created = SubscriptionQRCode.objects.get_or_create(
            subscription=subscription
        )
        
        # Si le QR Code existe déjà et a une image, on le retourne directement
        if not created and qr_code.qr_code_image:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'qr_code_url': qr_code.qr_code_image.url,
                    'user_name': subscription.user.get_full_name(),
                    'user_phone': subscription.user.phone,
                    'plan_name': subscription.plan.name,
                    'start_date': subscription.start_date.strftime('%d/%m/%Y'),
                    'is_new': False  # Indique que c'est un QR existant
                })
            else:
                messages.info(request, f"QR Code déjà existant pour {subscription.user.get_full_name()}.")
                return redirect('gestion_abonnements')
        
        # Générer un nouveau QR Code seulement si il n'existe pas
        qr_data = {
            'subscription_id': str(subscription.id),
            'user_name': subscription.user.get_full_name(),
            'user_phone': subscription.user.phone,
            'plan_name': subscription.plan.name,
            'start_date': subscription.start_date.isoformat(),
            'company': 'ECOCITY'
        }
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        # Créer l'image
        img = qr.make_image(fill_color="#2E8B57", back_color="white")
        
        # Sauvegarder l'image
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        # Sauvegarder dans le modèle
        filename = f"qr_code_{subscription.id}.png"
        qr_code.qr_code_image.save(filename, ContentFile(buffer.read()), save=True)
        qr_code.save()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'qr_code_url': qr_code.qr_code_image.url,
                'user_name': subscription.user.get_full_name(),
                'user_phone': subscription.user.phone,
                'plan_name': subscription.plan.name,
                'start_date': subscription.start_date.strftime('%d/%m/%Y'),
                'is_new': True  # Indique que c'est un nouveau QR
            })
        else:
            messages.success(request, f"QR Code généré avec succès pour {subscription.user.get_full_name()}.")
            return redirect('gestion_abonnements')
            
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            print('ici la faute est la')
            return JsonResponse({'success': False, 'error': str(e)})
        else:
            print(e)
            messages.error(request, f"Erreur lors de la génération du QR Code: {str(e)}")
            return redirect('gestion_abonnements')
        

@login_required(login_url='login')
@user_passes_test(is_admin)
def telecharger_qr_code(request, subscription_id):
    """Télécharger le QR Code"""
    try:
        subscription = get_object_or_404(Subscription, id=subscription_id)
        qr_code = get_object_or_404(SubscriptionQRCode, subscription=subscription)
        
        if qr_code.qr_code_image:
            response = HttpResponse(qr_code.qr_code_image.read(), content_type='image/png')
            response['Content-Disposition'] = f'attachment; filename="QRCode_ECOCITY_{subscription.user.get_full_name().replace(" ", "_")}.png"'
            return response
        else:
            messages.error(request, "Aucun QR Code trouvé pour cet abonnement.")
            return redirect('gestion_abonnements')
            
    except Exception as e:
        messages.error(request, f"Erreur lors du téléchargement: {str(e)}")
        return redirect('gestion_abonnements')
    

@login_required(login_url='login')
@user_passes_test(is_admin)
def recuperer_qr_code(request, subscription_id):
    """Récupérer un QR Code existant depuis la base de données"""
    try:
        subscription = get_object_or_404(Subscription, id=subscription_id)
        
        # Vérifier si un QR Code existe
        try:
            qr_code = SubscriptionQRCode.objects.get(subscription=subscription)
            
            if qr_code.qr_code_image :
                return JsonResponse({
                    'success': True,
                    'qr_code_url': qr_code.qr_code_image.url,
                    'user_name': subscription.user.get_full_name(),
                    'user_phone': subscription.user.phone,
                    'plan_name': subscription.plan.name,
                    'start_date': subscription.start_date.strftime('%d/%m/%Y'),
                    'exists': True
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'QR Code existe mais aucune image trouvée',
                    'exists': True
                })
                
        except SubscriptionQRCode.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Aucun QR Code trouvé pour cet abonnement',
                'exists': False
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'exists': False
        })

@login_required(login_url='login')
@user_passes_test(is_admin)
def detail_utilisateur(request, user_id):
    """Page de détail d'un utilisateur"""
    user = get_object_or_404(CustomUser, id=user_id)
    subscriptions = user.subscriptions.select_related('plan', 'zone').prefetch_related('collection_days')
    payments = Payment.objects.filter(subscription__user=user).order_by('-payment_date')
    
    context = {
        'user_detail': user,
        'subscriptions': subscriptions,
        'payments': payments,
    }
    
    return render(request, 'administrations/detail_utilisateur.html', context)

@login_required(login_url='login')
@user_passes_test(is_admin)
def detail_abonnement(request, subscription_id):
    """Page de détail d'un abonnement"""
    subscription = get_object_or_404(
        Subscription.objects.select_related('user', 'plan', 'zone', 'address'),
        id=subscription_id
    )
    
    # Précharger le QR Code s'il existe
    try:
        qr_code = subscription.qr_code
    except SubscriptionQRCode.DoesNotExist:
        qr_code = None
    
    collection_days = subscription.collection_days.select_related('day')
    collection_requests = subscription.collections.order_by('-scheduled_date')
    payments = subscription.payments.order_by('-payment_date')
    
    context = {
        'subscription_detail': subscription,
        'qr_code': qr_code,
        'collection_days': collection_days,
        'collection_requests': collection_requests,
        'payments': payments,
    }
    
    return render(request, 'administrations/detail_abonnement.html', context)

# API endpoints pour AJAX
@login_required(login_url='login')
@user_passes_test(is_admin)
def api_utilisateurs(request):
    """API pour les données utilisateurs"""
    search = request.GET.get('search', '')
    user_type = request.GET.get('user_type', '')
    
    users = CustomUser.objects.all()
    
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search)
        )
    
    if user_type:
        users = users.filter(user_type=user_type)
    
    users_data = []
    for user in users[:50]:  # Limiter à 50 résultats
        users_data.append({
            'id': str(user.id),
            'full_name': user.get_full_name(),
            'email': user.email,
            'phone': user.phone,
            'user_type': user.get_user_type_display(),
            'city': user.city.city if user.city else '',
            'date_joined': user.date_joined.strftime('%d/%m/%Y'),
            'is_verified': user.is_verified,
        })
    
    return JsonResponse({'users': users_data})

@login_required(login_url='login')
@user_passes_test(is_admin)
def api_abonnements(request):
    """API pour les données abonnements"""
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    subscriptions = Subscription.objects.select_related('user', 'plan', 'zone')
    
    if status:
        subscriptions = subscriptions.filter(status=status)
    
    if search:
        subscriptions = subscriptions.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(plan__name__icontains=search)
        )
    
    subscriptions_data = []
    for subscription in subscriptions[:50]:  # Limiter à 50 résultats
        subscriptions_data.append({
            'id': str(subscription.id),
            'user_name': subscription.user.get_full_name(),
            'user_phone': subscription.user.phone,
            'plan_name': subscription.plan.name,
            'plan_type': subscription.plan.get_plan_type_display(),
            'status': subscription.get_status_display(),
            'start_date': subscription.start_date.strftime('%d/%m/%Y'),
            'end_date': subscription.end_date.strftime('%d/%m/%Y') if subscription.end_date else '',
            'zone': subscription.zone.nom if subscription.zone else '',
            'price': float(subscription.custom_price or subscription.plan.price),
        })
    
    return JsonResponse({'subscriptions': subscriptions_data})

@login_required(login_url='login')
@user_passes_test(is_admin)
def api_statistiques(request):
    """API pour les statistiques en temps réel"""
    # Statistiques des abonnements
    subscription_stats = Subscription.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(status='active')),
        suspended=Count('id', filter=Q(status='suspended')),
        inactive=Count('id', filter=Q(status='inactive'))
    )
    
    # Statistiques des utilisateurs
    user_stats = CustomUser.objects.aggregate(
        total=Count('id'),
        clients=Count('id', filter=Q(user_type='client')),
        collecteurs=Count('id', filter=Q(user_type='collecteur')),
        admins=Count('id', filter=Q(user_type='admin'))
    )
    
    # Chiffre d'affaires du mois
    current_month_revenue = Payment.objects.filter(
        status='completed',
        payment_date__month=timezone.now().month,
        payment_date__year=timezone.now().year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    return JsonResponse({
        'subscriptions': subscription_stats,
        'users': user_stats,
        'current_month_revenue': float(current_month_revenue),
        'last_updated': timezone.now().strftime('%d/%m/%Y %H:%M:%S')
    })

@login_required(login_url='login')
@user_passes_test(is_admin)
def gestion_collecte(request):
    """Gestion de la collecte des déchets"""
    # Implémentation de la gestion de collecte
    return render(request, 'administrations/gestion_collecte.html')


# les vues pour recuperer les informations des client dont l'abonnement expire dans 8 jour
@staff_member_required
def export_abonnements_expirant(request):
    """
    Exporte les clients dont l'abonnement expire dans X jours
    Format: Fichier XLS sans entêtes, juste 2 colonnes (Nom | Téléphone)
    """
    # Récupérer les paramètres
    days = request.GET.get('days')
    include_all = request.GET.get('include_all')
    
    today = timezone.now().date()
    abonnements = Subscription.objects.filter(
        status='active',
        end_date__isnull=False
    ).select_related('user')
    
    if include_all and include_all == '1':
        # Inclure tous les abonnements actifs
        subscriptions = abonnements
        filename = f"tous_abonnements_{today.strftime('%Y%m%d')}.xlsx"
    else:
        try:
            days = int(days) if days else 7
            if days < 1:
                days = 1
            elif days > 365:
                days = 365
        except (ValueError, TypeError):
            days = 7
        
        # Calculer la date limite
        limit_date = today + timedelta(days=days)
        
        # Filtrer les abonnements qui expirent dans X jours
        subscriptions = abonnements.filter(
            end_date__lte=limit_date,
            end_date__gte=today
        )
        filename = f"abonnements_expirant_{days}_jours_{today.strftime('%Y%m%d')}.xlsx"
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Remplir les données (sans entêtes)
    row_num = 1
    for subscription in subscriptions:
        client = subscription.user
        nom_complet = f"{client.last_name} {client.first_name}".strip()
        if not nom_complet:
            nom_complet = client.username or "Client sans nom"
        
        # Colonne A: Nom complet
        ws.cell(row=row_num, column=1, value=nom_complet)
        
        # Colonne B: Numéro de téléphone
        phone = client.phone if client.phone else ""
        ws.cell(row=row_num, column=2, value=str(phone))
        
        row_num += 1
    
    # Ajuster la largeur des colonnes
    for col in range(1, 3):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 30
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    wb.save(response)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le workbook dans la réponse
    
    return response
        

@staff_member_required
def get_abonnements_expirant_stats(request):
    """
    Vue AJAX pour récupérer les statistiques des abonnements expirant
    """
    try:
        date_limite = timezone.now().date() + timedelta(days=8)
        aujourd_hui = timezone.now().date()
        
        # Statistiques générales
        total_expirant = Subscription.objects.filter(
            status='active',
            end_date__isnull=False,
            end_date__gte=aujourd_hui,
            end_date__lte=date_limite
        ).count()
        
        # Statistiques par jour
        jours_stats = []
        for i in range(1, 9):
            date = aujourd_hui + timedelta(days=i)
            count = Subscription.objects.filter(
                status='active',
                end_date=date
            ).count()
            
            if count > 0:
                jours_stats.append({
                    'date': date.strftime('%d/%m/%Y'),
                    'count': count,
                    'jour_semaine': date.strftime('%A')
                })
        
        return JsonResponse({
            'success': True,
            'total': total_expirant,
            'date_limite': date_limite.strftime('%d/%m/%Y'),
            'date_aujourd_hui': aujourd_hui.strftime('%d/%m/%Y'),
            'par_jour': jours_stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# les vues pour recuperer les clients inactif : 

@staff_member_required
def export_clients_inactifs_separe(request):
    """
    Exporte les clients avec abonnements inactifs dans deux fichiers Excel séparés
    - Un fichier pour les numéros Orange Cameroun (préfixes 65, 69, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99)
    - Un fichier pour les numéros MTN Cameroun (préfixes 67, 68, 6, 2)
    Format: 2 colonnes (Nom, Téléphone) sans en-têtes
    """
    
    # Récupérer tous les utilisateurs avec abonnements inactifs
    # Un client est considéré inactif si tous ses abonnements sont inactifs
    utilisateurs = CustomUser.objects.filter(
        user_type='client',
        is_active=True
    ).prefetch_related('subscriptions')
    
    # Filtrer pour n'avoir que les clients avec abonnements inactifs
    clients_inactifs = []
    for user in utilisateurs:
        # Vérifier si l'utilisateur a des abonnements
        abonnements = user.subscriptions.all()
        if abonnements.exists():
            # Si tous les abonnements sont inactifs, on inclut le client
            if all(sub.status != 'active' for sub in abonnements):
                if user.phone:  # S'assurer que le numéro existe
                    clients_inactifs.append({
                        'nom': f"{user.first_name} {user.last_name}".strip() or user.username,
                        'telephone': str(user.phone)
                    })
        else:
            # Si l'utilisateur n'a aucun abonnement, on l'inclut aussi
            if user.phone:
                clients_inactifs.append({
                    'nom': f"{user.first_name} {user.last_name}".strip() or user.username,
                    'telephone': str(user.phone)
                })
    
    # Fonction pour détecter l'opérateur camerounais
    def get_operateur(telephone):
        # Nettoyer le numéro (enlever les espaces, +, etc.)
        numero = re.sub(r'[\s\+\-\(\)]', '', telephone)
        
        # Si le numéro commence par 237 (indicatif Cameroun), on le retire
        if numero.startswith('237'):
            numero = numero[3:]
        
        # Si le numéro fait moins de 9 chiffres, on le complète (au cas où)
        if len(numero) < 9:
            return 'inconnu'
        
        # Prendre les 2 premiers chiffres après l'indicatif éventuel
        prefix = numero[:2] if len(numero) >= 2 else numero
        
        # Orange Cameroun : 65, 69, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99
        orange_prefixes = ['40', '55', '56', '57', '58', '59', '86', '87', '88', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', '99']
        
        # MTN Cameroun : 67, 68, 62, 63, 64, 65? (à vérifier) mais aussi 6, 2
        mtn_prefixes = ['50', '51', '52', '53', '54', '70', '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', '81', '82', '83']
        
        if prefix in orange_prefixes:
            return 'orange'
        elif prefix in mtn_prefixes or prefix.startswith('6') or prefix.startswith('2'):
            return 'mtn'
        else:
            return 'inconnu'
    
    # Séparer les clients par opérateur
    orange_clients = []
    mtn_clients = []
    
    for client in clients_inactifs:
        operateur = get_operateur(client['telephone'])
        if operateur == 'orange':
            orange_clients.append([client['nom'], client['telephone']])
        elif operateur == 'mtn':
            mtn_clients.append([client['nom'], client['telephone']])
        # On ignore les opérateurs inconnus
    
    # Créer le fichier pour Orange Cameroun
    wb_orange = openpyxl.Workbook()
    ws_orange = wb_orange.active
    
    # Ajouter les données sans en-têtes
    for row in orange_clients:
        ws_orange.append(row)
    
    # Ajuster la largeur des colonnes
    ws_orange.column_dimensions['A'].width = 30
    ws_orange.column_dimensions['B'].width = 20
    
    # Créer le fichier pour MTN Cameroun
    wb_mtn = openpyxl.Workbook()
    ws_mtn = wb_mtn.active
    
    # Ajouter les données sans en-têtes
    for row in mtn_clients:
        ws_mtn.append(row)
    
    # Ajuster la largeur des colonnes
    ws_mtn.column_dimensions['A'].width = 30
    ws_mtn.column_dimensions['B'].width = 20
    
    # Créer un fichier ZIP contenant les deux fichiers Excel
    from io import BytesIO
    import zipfile
    
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Sauvegarder le fichier Orange dans le ZIP
        orange_buffer = BytesIO()
        wb_orange.save(orange_buffer)
        orange_buffer.seek(0)
        zip_file.writestr('clients_inactifs_orange.xlsx', orange_buffer.getvalue())
        
        # Sauvegarder le fichier MTN dans le ZIP
        mtn_buffer = BytesIO()
        wb_mtn.save(mtn_buffer)
        mtn_buffer.seek(0)
        zip_file.writestr('clients_inactifs_mtn.xlsx', mtn_buffer.getvalue())
    
    zip_buffer.seek(0)
    
    # Générer le nom du fichier avec la date
    date_str = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'clients_inactifs_operateurs_{date_str}.zip'
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@staff_member_required
def export_clients_inactifs_orange(request):
    """
    Exporte uniquement les clients Orange Cameroun avec abonnements inactifs
    """
    return export_clients_inactifs_operateur(request, 'orange')


@staff_member_required
def export_clients_inactifs_mtn(request):
    """
    Exporte uniquement les clients MTN Cameroun avec abonnements inactifs
    """
    return export_clients_inactifs_operateur(request, 'mtn')


def export_clients_inactifs_operateur(request, operateur):
    """
    Fonction générique pour exporter les clients d'un opérateur spécifique
    """
    import re
    
    # Récupérer tous les utilisateurs avec abonnements inactifs
    utilisateurs = CustomUser.objects.filter(
        user_type='client',
        is_active=True
    ).prefetch_related('subscriptions')
    
    clients_inactifs = []
    for user in utilisateurs:
        abonnements = user.subscriptions.all()
        
        # Inclure si pas d'abonnement ou tous inactifs
        if not abonnements.exists() or all(sub.status != 'active' for sub in abonnements):
            if user.phone:
                clients_inactifs.append({
                    'nom': f"{user.first_name} {user.last_name}".strip() or user.username,
                    'telephone': str(user.phone)
                })
    
    # Filtrer par opérateur
    def is_orange(telephone):
        numero = re.sub(r'[\s\+\-\(\)]', '', telephone)
        if numero.startswith('237'):
            numero = numero[3:]
        prefix = numero[:2] if len(numero) >= 2 else numero
        orange_prefixes = ['40', '55', '56', '57', '58', '59', '86', '87', '88', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', '99']
        return prefix in orange_prefixes
    
    def is_mtn(telephone):
        numero = re.sub(r'[\s\+\-\(\)]', '', telephone)
        if numero.startswith('237'):
            numero = numero[3:]
        prefix = numero[:2] if len(numero) >= 2 else numero
        mtn_prefixes = ['50', '51', '52', '53', '54', '70', '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', '81', '82', '83']
        return prefix in mtn_prefixes or prefix.startswith('6') or prefix.startswith('2')
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Ajouter les données
    count = 0
    for client in clients_inactifs:
        if operateur == 'orange' and is_orange(client['telephone']):
            ws.append([client['nom'], client['telephone']])
            count += 1
        elif operateur == 'mtn' and is_mtn(client['telephone']):
            ws.append([client['nom'], client['telephone']])
            count += 1
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Si aucun client trouvé, ajouter un message
    if count == 0:
        ws.append(["Aucun client trouvé", ""])
    
    # Générer la réponse
    date_str = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'clients_inactifs_{operateur}_{date_str}.xlsx'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

