# app/gaz_admin_views.py
import uuid

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.contrib import messages
from django.views.decorators.http import require_POST
import json
import os

from django.core.files.storage import default_storage
from .models import (
    GasProduct, GasOrder, GasCylinder, GasInventory, 
    GasPromotion, GasDeliveryTracking, CustomUser, Tricycle, Zone, Notification, GasOrderItem
)

def is_admin(user):
    return user.is_authenticated and user.user_type == 'admin'

@login_required
@user_passes_test(is_admin)
def admin_gaz_dashboard(request):
    """Tableau de bord administration gaz"""
    
    # Statistiques du jour
    today = timezone.now().date()
    yesterday = today - timezone.timedelta(days=1)
    
    today_orders = GasOrder.objects.filter(order_date__date=today).count()
    yesterday_orders = GasOrder.objects.filter(order_date__date=yesterday).count()
    
    if yesterday_orders > 0:
        today_percentage = ((today_orders - yesterday_orders) / yesterday_orders) * 100
    else:
        today_percentage = 100 if today_orders > 0 else 0
    
    # Commandes par statut
    pending_orders = GasOrder.objects.filter(status='pending').count()
    in_transit_orders = GasOrder.objects.filter(status='in_transit').count()
    
    # Produits avec stock faible
    low_stock_count = GasProduct.objects.filter(
        stock_available__lte=5,
        is_available=True
    ).count()
    
    # Commandes récentes
    recent_orders = GasOrder.objects.select_related(
        'customer', 'assigned_collector'
    ).prefetch_related('items__product').order_by('-order_date')[:10]
    
    # Produits
    products = GasProduct.objects.filter(is_available=True).order_by('gaz_type', 'size_kg')
    
    # Statistiques bouteilles
    cylinders_stats = {
        'available': GasCylinder.objects.filter(status='available').count(),
        'with_customer': GasCylinder.objects.filter(status='with_customer').count(),
        'in_transit': GasCylinder.objects.filter(status='in_transit').count(),
        'damaged': GasCylinder.objects.filter(status='damaged').count(),
    }
    
    # Dernières livraisons
    delivered_orders = GasOrder.objects.filter(
        status='delivered'
    ).select_related(
        'customer', 'assigned_collector'
    ).order_by('-delivery_date')[:5]
    
    stats = {
        'today_orders': today_orders,
        'today_percentage': round(today_percentage, 1),
        'pending_orders': pending_orders,
        'in_transit_orders': in_transit_orders,
        'low_stock_count': low_stock_count,
    }
    
    context = {
        'stats': stats,
        'recent_orders': recent_orders,
        'products': products,
        'cylinders_stats': cylinders_stats,
        'delivered_orders': delivered_orders,
    }
    
    return render(request, 'admin_gaz/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def admin_gaz_products(request):
    """Gestion des produits gaz"""
    
    # Filtres
    products = GasProduct.objects.all().order_by('gaz_type', 'size_kg')
    
    if request.GET.get('gaz_type'):
        products = products.filter(gaz_type=request.GET['gaz_type'])
    
    if request.GET.get('size'):
        products = products.filter(size_kg=int(request.GET['size']))
    
    if request.GET.get('is_available') == 'true':
        products = products.filter(is_available=True)
    elif request.GET.get('is_available') == 'false':
        products = products.filter(is_available=False)
    
    # Pagination
    paginator = Paginator(products, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'products': page_obj,
        'gaz_types': GasProduct.GAZ_TYPE_CHOICES,
        'size_choices': GasProduct.SIZE_CHOICES,
    }
    
    return render(request, 'admin_gaz/products.html', context)

@login_required
@user_passes_test(is_admin)
def admin_gaz_orders(request):
    """Gestion des commandes de gaz"""
    
    orders = GasOrder.objects.select_related(
        'customer', 'assigned_collector', 'zone'
    ).prefetch_related('items__product').all().order_by('-order_date')
    
    # Filtres
    if request.GET.get('order_number'):
        orders = orders.filter(order_number__icontains=request.GET['order_number'])
    
    if request.GET.get('customer'):
        orders = orders.filter(
            Q(customer__first_name__icontains=request.GET['customer']) |
            Q(customer__last_name__icontains=request.GET['customer']) |
            Q(customer__username__icontains=request.GET['customer'])
        )
    
    if request.GET.get('status'):
        orders = orders.filter(status=request.GET['status'])
    
    if request.GET.get('payment_status'):
        orders = orders.filter(payment_status=request.GET['payment_status'])
    
    if request.GET.get('date_from'):
        orders = orders.filter(order_date__date__gte=request.GET['date_from'])
    
    if request.GET.get('date_to'):
        orders = orders.filter(order_date__date__lte=request.GET['date_to'])
    
    # Totaux
    total_amount = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    delivered_count = orders.filter(status='delivered').count()
    pending_count = orders.filter(status='pending').count()
    
    # Pagination
    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Liste des livreurs disponibles
    available_collectors = CustomUser.objects.filter(
        user_type='collecteur',
        is_active=True
    ).annotate(
        active_deliveries=Count('assigned_gas_deliveries', 
                               filter=Q(assigned_gas_deliveries__status__in=['assigned', 'in_transit']))
    ).order_by('active_deliveries')
    
    available_tricycles = Tricycle.objects.filter(status='active')
    
    context = {
        'orders': page_obj,
        'total_amount': total_amount,
        'delivered_count': delivered_count,
        'pending_count': pending_count,
        'order_statuses': GasOrder.ORDER_STATUS_CHOICES,
        'payment_statuses': GasOrder.PAYMENT_STATUS_CHOICES,
        'available_collectors': available_collectors,
        'available_tricycles': available_tricycles,
    }
    
    return render(request, 'admin_gaz/orders.html', context)

@login_required
@user_passes_test(is_admin)
def admin_gaz_order_detail(request, order_id):
    """Détail d'une commande"""
    
    order = get_object_or_404(
        GasOrder.objects.select_related(
            'customer', 'address', 'assigned_collector', 'assigned_tricycle'
        ).prefetch_related(
            'items__product', 'tracking_updates', 'items__assigned_cylinders'
        ),
        id=order_id
    )
    
    if request.method == 'POST':
        # Mise à jour de la commande
        action = request.POST.get('action')
        
        if action == 'update_status':
            new_status = request.POST.get('status')
            order.status = new_status
            
            if new_status == 'confirmed':
                order.confirmation_date = timezone.now()
            elif new_status == 'preparing':
                order.preparation_date = timezone.now()
            elif new_status == 'in_transit':
                # Créer une entrée de suivi
                GasDeliveryTracking.objects.create(
                    order=order,
                    status='in_transit',
                    notes="Commande en cours de livraison"
                )
            elif new_status == 'delivered':
                order.delivery_date = timezone.now()
            
            order.save()
            messages.success(request, "Statut mis à jour avec succès")
            
        elif action == 'assign_collector':
            collector_id = request.POST.get('collector_id')
            tricycle_id = request.POST.get('tricycle_id')
            
            if collector_id:
                collector = get_object_or_404(CustomUser, id=collector_id)
                order.assigned_collector = collector
                
                if tricycle_id:
                    tricycle = get_object_or_404(Tricycle, id=tricycle_id)
                    order.assigned_tricycle = tricycle
                
                order.status = 'assigned'
                order.save()
                messages.success(request, "Livreur assigné avec succès")
        
        return redirect('admin_gaz_order_detail', order_id=order.id)
    
    context = {
        'order': order,
        'order_statuses': GasOrder.ORDER_STATUS_CHOICES,
        'available_collectors': CustomUser.objects.filter(
            user_type='collecteur', is_active=True
        ),
        'tracking_updates': order.tracking_updates.all().order_by('-created_at'),
    }
    
    return render(request, 'admin_gaz/order_detail.html', context)

@login_required
@user_passes_test(is_admin)
@require_POST
def admin_gaz_create_order(request):
    """Créer une nouvelle commande manuellement"""
    
    try:
        data = json.loads(request.body)
        
        # Créer la commande
        order = GasOrder.objects.create(
            customer_id=data['customer_id'],
            address_id=data['address_id'],
            delivery_type=data.get('delivery_type', 'standard'),
            scheduled_date=data.get('scheduled_date'),
            special_instructions=data.get('instructions', ''),
            payment_method=data.get('payment_method', 'cash'),
            payment_status='pending'
        )
        
        # Ajouter les articles
        for item in data['items']:
            product = GasProduct.objects.get(id=item['product_id'])
            order_item = order.items.create(
                product=product,
                quantity=item['quantity'],
                unit_price=product.price,
                unit_deposit=product.deposit_price
            )
            
            # Mettre à jour le stock
            product.stock_available -= item['quantity']
            product.save()
        
        # Calculer les totaux
        order.calculate_totals()
        
        return JsonResponse({
            'success': True,
            'order_id': str(order.id),
            'order_number': order.order_number
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@user_passes_test(is_admin)
def admin_gaz_cylinders(request):
    """Gestion des bouteilles de gaz"""
    
    cylinders = GasCylinder.objects.select_related(
        'product', 'current_owner'
    ).all().order_by('-created_at')
    
    # Filtres
    if request.GET.get('status'):
        cylinders = cylinders.filter(status=request.GET['status'])
    
    if request.GET.get('product'):
        cylinders = cylinders.filter(product_id=request.GET['product'])
    
    # Pagination
    paginator = Paginator(cylinders, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'cylinders': page_obj,
        'products': GasProduct.objects.all(),
        'status_choices': GasCylinder.STATUS_CHOICES,
    }
    
    return render(request, 'admin_gaz/cylinders.html', context)

@login_required
@user_passes_test(is_admin)
def admin_gaz_inventory(request):
    """Gestion des stocks"""
    
    inventory = GasInventory.objects.select_related(
        'product', 'tricycle'
    ).all().order_by('location_type', 'product')
    
    # Résumé par produit
    product_summary = GasProduct.objects.annotate(
        total_stock=Sum('inventory_items__quantity_available'),
        total_reserved=Sum('inventory_items__quantity_reserved'),
        warehouse_count=Count('inventory_items', 
                             filter=Q(inventory_items__location_type='warehouse')),
        tricycle_count=Count('inventory_items',
                           filter=Q(inventory_items__location_type='tricycle'))
    ).filter(is_available=True)
    
    if request.method == 'POST':
        # Mise à jour d'inventaire
        inventory_id = request.POST.get('inventory_id')
        new_quantity = int(request.POST.get('quantity'))
        adjustment_type = request.POST.get('adjustment_type')
        reason = request.POST.get('reason', '')
        
        inv = get_object_or_404(GasInventory, id=inventory_id)
        
        if adjustment_type == 'add':
            inv.quantity_available += new_quantity
        elif adjustment_type == 'remove':
            if inv.quantity_available >= new_quantity:
                inv.quantity_available -= new_quantity
            else:
                messages.error(request, "Stock insuffisant")
                return redirect('admin_gaz_inventory')
        elif adjustment_type == 'set':
            inv.quantity_available = new_quantity
        
        inv.save()
        
        # Journaliser l'action
        # Ici vous pouvez ajouter un modèle d'historique si nécessaire
        
        messages.success(request, "Inventaire mis à jour avec succès")
        return redirect('admin_gaz_inventory')
    
    context = {
        'inventory': inventory,
        'product_summary': product_summary,
        'location_types': GasInventory.LOCATION_TYPE_CHOICES,
    }
    
    return render(request, 'admin_gaz/inventory.html', context)

@login_required
@user_passes_test(is_admin)
def admin_gaz_promotions(request):
    """Gestion des promotions"""
    
    promotions = GasPromotion.objects.all().order_by('-created_at')
    
    # Promotions actives
    active_promotions = promotions.filter(
        is_active=True,
        start_date__lte=timezone.now(),
        end_date__gte=timezone.now()
    )
    
    if request.method == 'POST':
        # Créer une nouvelle promotion
        promo = GasPromotion.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description'),
            promotion_type=request.POST.get('promotion_type'),
            discount_value=request.POST.get('discount_value') or None,
            discount_percentage=request.POST.get('discount_percentage') or None,
            minimum_purchase=request.POST.get('minimum_purchase', 1),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date'),
            promo_code=request.POST.get('promo_code') or None,
            usage_limit=request.POST.get('usage_limit', 0),
            is_active=True
        )
        
        # Ajouter les produits applicables
        product_ids = request.POST.getlist('products')
        if product_ids:
            promo.applicable_products.set(product_ids)
        
        # Ajouter les zones applicables
        zone_ids = request.POST.getlist('zones')
        if zone_ids:
            promo.applicable_zones.set(zone_ids)
        
        messages.success(request, "Promotion créée avec succès")
        return redirect('admin_gaz_promotions')
    
    context = {
        'promotions': promotions,
        'active_promotions': active_promotions,
        'promotion_types': GasPromotion.PROMOTION_TYPE_CHOICES,
        'products': GasProduct.objects.filter(is_available=True),
        'zones': Zone.objects.filter(is_active=True),
    }
    
    return render(request, 'admin_gaz/promotions.html', context)

@login_required
@user_passes_test(is_admin)
@require_POST
def admin_gaz_update_stock(request, product_id):
    """Mettre à jour le stock d'un produit"""
    
    try:
        product = get_object_or_404(GasProduct, id=product_id)
        adjustment_type = request.POST.get('adjustment_type')
        quantity = int(request.POST.get('quantity', 0))
        reason = request.POST.get('reason', '')
        
        if adjustment_type == 'add':
            product.stock_available += quantity
        elif adjustment_type == 'remove':
            if product.stock_available >= quantity:
                product.stock_available -= quantity
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Stock insuffisant'
                })
        elif adjustment_type == 'set':
            product.stock_available = quantity
        
        product.save()
        
        # Ici vous pouvez journaliser l'action
        
        return JsonResponse({
            'success': True,
            'new_stock': product.stock_available
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
@user_passes_test(is_admin)
@require_POST
def admin_gaz_assign_order(request, order_id):
    """Assigner un livreur à une commande"""
    
    try:
        order = get_object_or_404(GasOrder, id=order_id)
        collector_id = request.POST.get('collector_id')
        tricycle_id = request.POST.get('tricycle_id')
        
        collector = get_object_or_404(CustomUser, id=collector_id)
        order.assigned_collector = collector
        
        if tricycle_id:
            tricycle = get_object_or_404(Tricycle, id=tricycle_id)
            order.assigned_tricycle = tricycle
        
        order.status = 'assigned'
        order.save()
        
        # Créer une notification pour le collecteur
        from .models import Notification
        Notification.create_notification(
            user=collector,
            title="Nouvelle livraison assignée",
            message=f"Commande {order.order_number} - {order.customer.get_full_name()}",
            notification_type='info',
            related_object=order,
            action_url=f'/collector/gas/delivery/{order.id}/'
        )
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
@user_passes_test(is_admin)
def admin_gaz_export_orders(request):
    """Exporter les commandes au format Excel"""
    
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Récupérer les commandes filtrées
    orders = GasOrder.objects.select_related(
        'customer', 'assigned_collector'
    ).prefetch_related('items__product').all()
    
    # Appliquer les mêmes filtres que la vue orders
    if request.GET.get('order_number'):
        orders = orders.filter(order_number__icontains=request.GET['order_number'])
    
    if request.GET.get('status'):
        orders = orders.filter(status=request.GET['status'])
    
    if request.GET.get('date_from'):
        orders = orders.filter(order_date__date__gte=request.GET['date_from'])
    
    if request.GET.get('date_to'):
        orders = orders.filter(order_date__date__lte=request.GET['date_to'])
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes Gaz"
    
    # En-têtes
    headers = ['N° Commande', 'Date', 'Client', 'Téléphone', 'Produits', 
               'Montant', 'Statut', 'Paiement', 'Livreur', 'Date Livraison']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Données
    for row, order in enumerate(orders, 2):
        produits = ', '.join([f"{item.quantity}x {item.product.size_kg}kg" 
                             for item in order.items.all()])
        
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.order_date.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3, value=order.customer.get_full_name() or order.customer.username)
        ws.cell(row=row, column=4, value=order.customer.phone)
        ws.cell(row=row, column=5, value=produits)
        ws.cell(row=row, column=6, value=float(order.total_amount))
        ws.cell(row=row, column=7, value=order.get_status_display())
        ws.cell(row=row, column=8, value=order.get_payment_status_display())
        ws.cell(row=row, column=9, value=order.assigned_collector.get_full_name() if order.assigned_collector else 'Non assigné')
        ws.cell(row=row, column=10, value=order.delivery_date.strftime('%d/%m/%Y %H:%M') if order.delivery_date else '')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=commandes_gaz_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

# API endpoints pour AJAX
@login_required
def api_gas_product_detail(request, product_id):
    """API pour obtenir les détails d'un produit"""
    
    product = get_object_or_404(GasProduct, id=product_id)
    
    data = {
        'id': str(product.id),
        'name': product.name,
        'gaz_type': product.gaz_type,
        'size_kg': product.size_kg,
        'price': float(product.price),
        'deposit_price': float(product.deposit_price),
        'stock_available': product.stock_available,
        'description': product.description,
        'is_available': product.is_available,
        'is_featured': product.is_featured,
    }
    
    return JsonResponse(data)




# app/gaz_admin_views.py


@login_required
@user_passes_test(is_admin)
def admin_gaz_add_product_ajax(request):
    """Vue AJAX pour ajouter un produit"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            name = request.POST.get('name')
            gaz_type = request.POST.get('gaz_type')
            size_kg = request.POST.get('size_kg')
            price = request.POST.get('price')
            deposit_price = request.POST.get('deposit_price')
            stock_available = request.POST.get('stock_available', 0)
            inventory_threshold = request.POST.get('inventory_threshold', 5)
            description = request.POST.get('description', '')
            is_available = request.POST.get('is_available') == 'on'
            is_featured = request.POST.get('is_featured') == 'on'
            
            # Validation
            if not all([name, gaz_type, size_kg, price, deposit_price]):
                return JsonResponse({
                    'success': False,
                    'message': 'Tous les champs obligatoires doivent être remplis'
                })
            
           
            
            
            # Créer le produit
            product = GasProduct(
                name=name,
                gaz_type=gaz_type,
                size_kg=size_kg,
                price=price,
                deposit_price=deposit_price,
                stock_available=stock_available,
                description=description,
                is_available=is_available,
                is_featured=is_featured
            )
            
            # Gérer l'upload d'image
            if 'image' in request.FILES:
                image = request.FILES['image']
                
                # Vérifier la taille (2MB max)
                if image.size > 2 * 1024 * 1024:
                    return JsonResponse({
                        'success': False,
                        'message': 'L\'image ne doit pas dépasser 2MB'
                    })
                
                # Vérifier le type
                if not image.content_type.startswith('image/'):
                    return JsonResponse({
                        'success': False,
                        'message': 'Le fichier doit être une image'
                    })
                
                # Générer un nom de fichier unique
                ext = os.path.splitext(image.name)[1]
                filename = f"gaz_products/{uuid.uuid4()}{ext}"
                
                # Sauvegarder l'image
                product.image.save(filename,image, save=False)
            
            product.save()
            
            
            
            # Préparer la réponse
            response_data = {
                'success': True,
                'product': {
                    'id': str(product.id),
                    'name': product.name,
                    'gaz_type': product.gaz_type,
                    'size_kg': product.size_kg,
                    'price': float(product.price),
                    'deposit_price': float(product.deposit_price),
                    'stock_available': product.stock_available,
                    'description': product.description,
                    'is_available': product.is_available,
                    'is_featured': product.is_featured,
                    'image_url': product.image.url if product.image else None
                }
            }
            
            return JsonResponse(response_data)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
@user_passes_test(is_admin)
def admin_gaz_update_product(request, product_id):
    """Vue AJAX pour mettre à jour un produit"""
    if request.method == 'POST':
        try:
            product = get_object_or_404(GasProduct, id=product_id)
            
            # Mettre à jour les champs
            product.name = request.POST.get('name', product.name)
            product.gaz_type = request.POST.get('gaz_type', product.gaz_type)
            product.size_kg = request.POST.get('size_kg', product.size_kg)
            product.price = request.POST.get('price', product.price)
            product.deposit_price = request.POST.get('deposit_price', product.deposit_price)
            product.stock_available = request.POST.get('stock_available', product.stock_available)
            product.description = request.POST.get('description', product.description)
            product.is_available = request.POST.get('is_available') == 'on'
            product.is_featured = request.POST.get('is_featured') == 'on'
            
            # Gérer l'upload d'image
            if 'image' in request.FILES:
                # Supprimer l'ancienne image si elle existe
                if product.image:
                    default_storage.delete(product.image.path)
                
                image = request.FILES['image']
                
                # Vérifier la taille
                if image.size > 2 * 1024 * 1024:
                    return JsonResponse({
                        'success': False,
                        'message': 'L\'image ne doit pas dépasser 2MB'
                    })
                
                # Générer un nom de fichier unique
                ext = os.path.splitext(image.name)[1]
                filename = f"gaz_products/{uuid.uuid4()}{ext}"
                
                # Sauvegarder la nouvelle image
                product.image.save(filename, image, save=False)
            
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Produit modifié avec succès'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
@user_passes_test(is_admin)
def admin_gaz_delete_product(request, product_id):
    """Vue AJAX pour supprimer un produit"""
    if request.method == 'POST':
        try:
            product = get_object_or_404(GasProduct, id=product_id)
            
            # Vérifier si le produit a des commandes associées
            if GasOrderItem.objects.filter(product=product).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Ce produit ne peut pas être supprimé car il a des commandes associées'
                })
            
            # Supprimer l'image si elle existe
            if product.image:
                default_storage.delete(product.image.path)
            
            # Supprimer le produit
            product.delete()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
@user_passes_test(is_admin)
def admin_gaz_update_stock(request, product_id):
    """Vue AJAX pour mettre à jour le stock"""
    if request.method == 'POST':
        try:
            product = get_object_or_404(GasProduct, id=product_id)
            adjustment_type = request.POST.get('adjustment_type')
            quantity = int(request.POST.get('quantity', 0))
            reason = request.POST.get('reason', '')
            
            old_stock = product.stock_available
            
            if adjustment_type == 'add':
                product.stock_available += quantity
            elif adjustment_type == 'remove':
                if product.stock_available >= quantity:
                    product.stock_available -= quantity
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Stock insuffisant'
                    })
            elif adjustment_type == 'set':
                product.stock_available = quantity
            
            product.save()
            
            # Journaliser le mouvement
            # Vous pouvez ajouter un modèle StockMovement ici si nécessaire
            
            # Créer une notification si le stock est faible
            if product.stock_available <= 5:
                Notification.create_notification(
                    user=request.user,
                    title="Stock faible",
                    message=f"Le produit {product.name} a un stock faible ({product.stock_available} unités)",
                    notification_type='warning'
                )
            
            return JsonResponse({
                'success': True,
                'new_stock': product.stock_available,
                'message': f'Stock mis à jour: {old_stock} → {product.stock_available}'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})